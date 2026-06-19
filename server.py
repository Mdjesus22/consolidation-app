"""
Reconciliation Tool — Flask Backend
Handles large file uploads, pandas-based reconciliation, and Excel download.
Start with: python server.py  (or double-click start.bat)
"""

from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
import pandas as pd
import io, os, uuid, threading, json, traceback
from collections import defaultdict

app = Flask(__name__, static_folder=None)
CORS(app)

# ── In-memory stores ──────────────────────────────────────────────────────────
file_store = {}   # file_id -> {bytes, filename, sheets, current_sheet, file_size}
job_store  = {}   # job_id  -> {status, progress, label, results, error, full_results_N, df_a_N, df_b_N, config}
_store_lock = threading.Lock()


# ── Serve the HTML app ────────────────────────────────────────────────────────
HTML_FILE = os.path.join(os.path.dirname(__file__),
                         'Reconciliation_tool (Snowflake vs Seller Center).html')

@app.route('/')
def index():
    return send_file(HTML_FILE)


# ── Normalisation (mirrors JS norm() logic) ───────────────────────────────────
def apply_norm(val, rules):
    """Apply normalisation rules to a single string value."""
    v = str(val) if val is not None else ''
    v = v.strip()
    rules_str = ' '.join(rules) if isinstance(rules, list) else str(rules)

    if 'lowercase' in rules_str:
        v = v.lower().strip()
    if 'round2' in rules_str:
        try:    v = f'{float(v):.2f}'
        except: pass
    if 'round0' in rules_str:
        try:    v = str(round(float(v)))
        except: pass
    if 'zero_blank' in rules_str:
        if v.strip().lower() in ('', '0', '0.0', 'nan', 'none', 'n/a', 'na', '-', 'null'):
            v = '__EMPTY__'
    return v


def build_key(row, side, keys):
    parts = []
    for k in keys:
        col  = k['colA'] if side == 'A' else k['colB']
        val  = str(row.get(col, '') or '').strip()
        parts.append(apply_norm(val, k.get('rules', [])))
    return '|||'.join(parts)


def merge_into(result_map, row, key, comps, side):
    """Insert row into map, summing numeric compare-columns on duplicate keys."""
    if key not in result_map:
        result_map[key] = {str(k): str(v) for k, v in row.items()}
    else:
        existing = result_map[key]
        for c in comps:
            col = c['colA'] if side == 'A' else c['colB']
            try:
                existing[col] = str(round(
                    (float(existing.get(col) or 0) + float(row.get(col) or 0)) * 1e6
                ) / 1e6)
            except Exception:
                pass   # non-numeric — keep first value


# ── Core reconciliation ───────────────────────────────────────────────────────
def reconcile_pair(df_a, df_b, keys, comps, progress_cb=None):
    rows_a = df_a.fillna('').to_dict('records')
    rows_b = df_b.fillna('').to_dict('records')
    total  = max(len(rows_a) + len(rows_b), 1)

    # Build map A
    map_a = {}
    for i, row in enumerate(rows_a):
        merge_into(map_a, row, build_key(row, 'A', keys), comps, 'A')
        if progress_cb and i % 10_000 == 0:
            progress_cb(int(i / total * 30) + 5,
                        f'Indexing File A… {i:,} / {len(rows_a):,}')

    # Build map B
    if progress_cb: progress_cb(38, 'Indexing File B…')
    map_b = {}
    for i, row in enumerate(rows_b):
        merge_into(map_b, row, build_key(row, 'B', keys), comps, 'B')
        if progress_cb and i % 10_000 == 0:
            progress_cb(int(i / total * 30) + 38,
                        f'Indexing File B… {i:,} / {len(rows_b):,}')

    # Compare
    if progress_cb: progress_cb(72, 'Comparing…')
    all_keys = set(map_a) | set(map_b)
    results  = []

    for ki, k in enumerate(all_keys):
        row_a = map_a.get(k)
        row_b = map_b.get(k)

        if row_a is None:
            results.append({'key': k, 'status': 'only_in_B',
                            'rowA': None, 'rowB': row_b, 'details': []})
        elif row_b is None:
            results.append({'key': k, 'status': 'only_in_A',
                            'rowA': row_a, 'rowB': None, 'details': []})
        else:
            ok      = True
            details = []
            for m in comps:
                va    = apply_norm(row_a.get(m['colA'], ''), m.get('rules', []))
                vb    = apply_norm(row_b.get(m['colB'], ''), m.get('rules', []))
                match = (va == '__EMPTY__' and vb == '__EMPTY__') or (va == vb)
                if not match:
                    ok = False
                details.append({
                    'colA': m['colA'], 'colB': m['colB'],
                    'va': row_a.get(m['colA'], ''),
                    'vb': row_b.get(m['colB'], ''),
                    'match': match,
                })
            results.append({'key': k,
                            'status': 'match' if ok else 'mismatch',
                            'rowA': row_a, 'rowB': row_b, 'details': details})

        if progress_cb and ki % 10_000 == 0:
            progress_cb(72 + int(ki / len(all_keys) * 23),
                        f'Comparing… {ki:,} / {len(all_keys):,}')

    return results


def compute_insights(results, keys, comps, date_col_a, date_col_b):
    """Pre-compute insight data server-side so the frontend doesn't need all rows."""
    bad = [r for r in results if r['status'] != 'match']

    # Issues by date
    date_counts = defaultdict(int)
    for r in bad:
        date_val = None
        if date_col_a and r.get('rowA'):
            date_val = r['rowA'].get(date_col_a)
        if not date_val and date_col_b and r.get('rowB'):
            date_val = r['rowB'].get(date_col_b)
        if date_val:
            d = str(date_val)[:10]
            date_counts[d] += 1

    # Mismatch by column
    col_counts = defaultdict(int)
    for r in results:
        if r['status'] == 'mismatch':
            for d in r['details']:
                if not d['match']:
                    lbl = d['colA'] if d['colA'] == d['colB'] else f"{d['colA']}/{d['colB']}"
                    col_counts[lbl] += 1

    mm_count    = sum(1 for r in results if r['status'] == 'mismatch')
    only_a      = sum(1 for r in results if r['status'] == 'only_in_A')
    only_b      = sum(1 for r in results if r['status'] == 'only_in_B')

    return {
        'byDate':      dict(sorted(date_counts.items(), reverse=True)[:30]),
        'byColumn':    dict(sorted(col_counts.items(), key=lambda x: -x[1])[:10]),
        'issueType':   {'Value mismatch': mm_count, 'Only in A': only_a, 'Only in B': only_b},
        'missingDir':  {'Only in A': only_a, 'Only in B': only_b},
    }


# ── Background job runner ─────────────────────────────────────────────────────
def run_job(job_id, config):
    pairs      = config['pairs']
    keys       = config['keys']
    comps      = config['comps']
    date_col_a = config.get('dateColA', '')
    date_col_b = config.get('dateColB', '')

    with _store_lock:
        job_store[job_id]['status'] = 'running'

    try:
        pair_results = []

        for pi, pair in enumerate(pairs):
            fid_a = pair['fileIdA']
            fid_b = pair['fileIdB']

            def progress_cb(pct, label, _pi=pi):
                overall = (_pi / len(pairs)) * 100 + pct / len(pairs)
                with _store_lock:
                    job_store[job_id]['progress'] = int(overall)
                    job_store[job_id]['label']    = f'Pair {_pi+1}/{len(pairs)}: {label}'

            progress_cb(0, 'Loading files…')

            fa = file_store[fid_a]
            fb = file_store[fid_b]

            def load_df(entry, sheet_override=None):
                entry['bytes'].seek(0)
                name  = entry['filename'].lower()
                sheet = sheet_override or entry.get('current_sheet')
                if name.endswith('.csv'):
                    return pd.read_csv(entry['bytes'], dtype=str)
                elif name.endswith('.tsv'):
                    return pd.read_csv(entry['bytes'], dtype=str, sep='\t')
                else:
                    return pd.read_excel(entry['bytes'], sheet_name=sheet, dtype=str)

            df_a = load_df(fa, pair.get('sheetA'))
            df_b = load_df(fb, pair.get('sheetB'))

            # Apply Group By if configured
            def apply_grouping(df, cfg):
                if not cfg or cfg.get('mode') != 'grouped':
                    return df
                gcol, scol = cfg.get('groupCol'), cfg.get('sumCol')
                # gcol can be a list (multi-column) or a single string
                if isinstance(gcol, str):
                    gcol = [gcol] if gcol else []
                gcol = [c for c in gcol if c in df.columns]  # drop any missing cols
                if not gcol or not scol or scol not in df.columns:
                    return df
                df = df.copy()
                df[scol] = pd.to_numeric(df[scol], errors='coerce').fillna(0)
                return df.groupby(gcol, as_index=False)[scol].sum()

            df_a = apply_grouping(df_a, pair.get('groupConfigA'))
            df_b = apply_grouping(df_b, pair.get('groupConfigB'))

            progress_cb(5, f'Loaded {len(df_a):,} + {len(df_b):,} rows — reconciling…')

            results = reconcile_pair(df_a, df_b, keys, comps, progress_cb)

            match    = sum(1 for r in results if r['status'] == 'match')
            mismatch = sum(1 for r in results if r['status'] == 'mismatch')
            only_a   = sum(1 for r in results if r['status'] == 'only_in_A')
            only_b   = sum(1 for r in results if r['status'] == 'only_in_B')

            # Limit rows sent to frontend (full set kept for download)
            non_match   = [r for r in results if r['status'] != 'match']
            match_sample= [r for r in results if r['status'] == 'match'][:200]
            display_rows = non_match[:2000] + match_sample

            insights = compute_insights(results, keys, comps, date_col_a, date_col_b)

            label = (f"{fa['filename'].rsplit('.',1)[0]} ↔ "
                     f"{fb['filename'].rsplit('.',1)[0]}")

            pair_results.append({
                'label':      label,
                'fileA':      fa['filename'],
                'fileB':      fb['filename'],
                'summary':    {'total': len(results), 'match': match,
                               'mismatch': mismatch,
                               'only_in_A': only_a, 'only_in_B': only_b},
                'displayRows': display_rows,
                'insights':   insights,
                'headersA':   list(df_a.columns),
                'headersB':   list(df_b.columns),
            })

            with _store_lock:
                job_store[job_id][f'full_results_{pi}'] = results
                job_store[job_id][f'df_a_{pi}']         = df_a
                job_store[job_id][f'df_b_{pi}']         = df_b

        with _store_lock:
            job_store[job_id].update({
                'status':   'done',
                'progress': 100,
                'label':    'Done!',
                'results':  pair_results,
                'config':   config,
            })

    except Exception as e:
        with _store_lock:
            job_store[job_id].update({
                'status': 'error',
                'error':  str(e),
                'trace':  traceback.format_exc(),
            })
        print(f'[job {job_id}] ERROR: {e}\n{traceback.format_exc()}')


# ── API routes ────────────────────────────────────────────────────────────────
@app.route('/api/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    f         = request.files['file']
    sheet_req = request.form.get('sheet')
    raw_bytes = f.read()
    file_id   = str(uuid.uuid4())
    buf       = io.BytesIO(raw_bytes)

    try:
        fname = f.filename.lower()
        if fname.endswith('.csv'):
            df     = pd.read_csv(buf, dtype=str)
            sheets = ['Sheet1']
            sheet  = 'Sheet1'
        elif fname.endswith('.tsv'):
            df     = pd.read_csv(buf, dtype=str, sep='\t')
            sheets = ['Sheet1']
            sheet  = 'Sheet1'
        else:
            buf.seek(0)
            xl     = pd.ExcelFile(buf)
            sheets = xl.sheet_names
            sheet  = sheet_req or sheets[0]
            buf.seek(0)
            df     = pd.read_excel(buf, sheet_name=sheet, dtype=str)

        with _store_lock:
            file_store[file_id] = {
                'bytes':         io.BytesIO(raw_bytes),
                'filename':      f.filename,
                'sheets':        sheets,
                'current_sheet': sheet,
                'file_size':     len(raw_bytes),
            }

        return jsonify({
            'fileId':       file_id,
            'filename':     f.filename,
            'sheets':       sheets,
            'currentSheet': sheet,
            'headers':      list(df.columns),
            'rowCount':     len(df),
            'fileSize':     len(raw_bytes),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sheet', methods=['POST'])
def change_sheet():
    data    = request.json or {}
    file_id = data.get('fileId')
    sheet   = data.get('sheet')

    if file_id not in file_store:
        return jsonify({'error': 'File not found'}), 404
    try:
        entry = file_store[file_id]
        entry['bytes'].seek(0)
        df = pd.read_excel(entry['bytes'], sheet_name=sheet, dtype=str)
        entry['current_sheet'] = sheet
        return jsonify({'headers': list(df.columns), 'rowCount': len(df)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/run', methods=['POST'])
def run():
    config = request.json or {}
    job_id = str(uuid.uuid4())
    with _store_lock:
        job_store[job_id] = {'status': 'pending', 'progress': 0,
                             'label': 'Queued…', 'results': None}
    t = threading.Thread(target=run_job, args=(job_id, config), daemon=True)
    t.start()
    return jsonify({'jobId': job_id})


@app.route('/api/job/<job_id>/status')
def job_status(job_id):
    if job_id not in job_store:
        return jsonify({'error': 'Not found'}), 404
    j = job_store[job_id]
    return jsonify({'status': j['status'], 'progress': j['progress'],
                    'label': j.get('label', ''), 'error': j.get('error')})


@app.route('/api/job/<job_id>/results')
def job_results(job_id):
    if job_id not in job_store:
        return jsonify({'error': 'Not found'}), 404
    j = job_store[job_id]
    if j['status'] != 'done':
        return jsonify({'error': 'Not done yet', 'status': j['status']}), 400
    return jsonify({'results': j['results']})


@app.route('/api/job/<job_id>/download')
def job_download(job_id):
    if job_id not in job_store:
        return jsonify({'error': 'Not found'}), 404
    j = job_store[job_id]
    if j['status'] != 'done':
        return jsonify({'error': 'Job not finished'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font

        config     = j.get('config', {})
        pairs      = config.get('pairs', [])
        keys       = config.get('keys', [])
        comps      = config.get('comps', [])
        shop_name  = request.args.get('shopName', '').strip()
        data_point = request.args.get('dataPoint', '').strip()

        HDR_FILL  = PatternFill('solid', fgColor='4F6EF7')
        HDR_FONT  = Font(color='FFFFFF', bold=True)
        MM_FILL   = PatternFill('solid', fgColor='FEE2E2')
        ONLY_FILL = PatternFill('solid', fgColor='FEF3C7')

        wb = Workbook()
        wb.remove(wb.active)

        for pi, pair in enumerate(pairs):
            full  = j.get(f'full_results_{pi}', [])
            df_a  = j.get(f'df_a_{pi}')
            df_b  = j.get(f'df_b_{pi}')
            base  = f'Pair{pi+1}'

            # Optional shop-name filter
            if shop_name:
                def has_shop(r):
                    row = r.get('rowA') or r.get('rowB') or {}
                    return any(shop_name.lower() in str(v).lower() for v in row.values())
                full = [r for r in full if has_shop(r)]

            mm_rows    = [r for r in full if r['status'] == 'mismatch']
            only_a_rows= [r for r in full if r['status'] == 'only_in_A']
            only_b_rows= [r for r in full if r['status'] == 'only_in_B']
            match_ct   = sum(1 for r in full if r['status'] == 'match')
            total_ct   = len(full)

            def write_sheet(ws, headers, rows, fill):
                ws.append(headers)
                for cell in ws[1]:
                    cell.fill, cell.font = HDR_FILL, HDR_FONT
                for row_data in rows:
                    ws.append(row_data)
                    for cell in ws[ws.max_row]:
                        cell.fill = fill

            # Mismatch sheet
            if mm_rows:
                ws = wb.create_sheet(f'{base}_mismatch')
                hdrs = (['STATUS','KEY'] +
                        [k['colA']+'_A' for k in keys] +
                        [k['colB']+'_B' for k in keys] +
                        [c['colA']+'_A' for c in comps] +
                        [c['colB']+'_B' for c in comps] +
                        ['DIFF_'+c['colA'] for c in comps])
                rows_data = []
                for r in mm_rows:
                    rd = (['mismatch', r['key']] +
                          [r['rowA'].get(k['colA'],'') if r['rowA'] else '' for k in keys] +
                          [r['rowB'].get(k['colB'],'') if r['rowB'] else '' for k in keys] +
                          [r['rowA'].get(c['colA'],'') if r['rowA'] else '' for c in comps] +
                          [r['rowB'].get(c['colB'],'') if r['rowB'] else '' for c in comps] +
                          ['' if d['match'] else f'⚠ {d["va"]} ≠ {d["vb"]}' for d in r['details']])
                    rows_data.append(rd)
                write_sheet(ws, hdrs, rows_data, MM_FILL)

            # Only in A
            if only_a_rows and df_a is not None:
                ws   = wb.create_sheet(f'{base}_onlyA')
                cols = list(df_a.columns)
                write_sheet(ws, ['STATUS']+cols,
                            [['only_in_A']+[r['rowA'].get(c,'') for c in cols]
                             for r in only_a_rows], ONLY_FILL)

            # Only in B
            if only_b_rows and df_b is not None:
                ws   = wb.create_sheet(f'{base}_onlyB')
                cols = list(df_b.columns)
                write_sheet(ws, ['STATUS']+cols,
                            [['only_in_B']+[r['rowB'].get(c,'') for c in cols]
                             for r in only_b_rows], ONLY_FILL)

            # Summary
            ws = wb.create_sheet(f'{base}_summary')
            ws.append(['Metric','Count'])
            for cell in ws[1]: cell.fill, cell.font = HDR_FILL, HDR_FONT
            for label, val in [
                ('Total', total_ct), ('Matched', match_ct),
                ('Mismatched', len(mm_rows)),
                ('Only in A', len(only_a_rows)),
                ('Only in B', len(only_b_rows)),
                ('Match %', f'{match_ct/total_ct*100:.1f}%' if total_ct else '0%'),
            ]:
                ws.append([label, val])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        fname = f'reconciliation_results{"_"+shop_name if shop_name else ""}.xlsx'
        return send_file(out,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=fname)
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'files': len(file_store), 'jobs': len(job_store)})


if __name__ == '__main__':
    import webbrowser, threading as _th
    port = 5050
    print(f'\n🚀  Reconciliation Tool server running at http://localhost:{port}')
    print('   Close this window to stop the server.\n')
    _th.Timer(1.2, lambda: webbrowser.open(f'http://localhost:{port}')).start()
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
